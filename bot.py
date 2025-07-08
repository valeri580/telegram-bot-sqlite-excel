import logging
import asyncio
from aiogram import Bot, Dispatcher, types, Router, F
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
from config import TELEGRAM_BOT_TOKEN, GIGACHAT_API_KEY, YANDEX_API_KEY
import sqlite3
import openpyxl
from datetime import datetime

logging.basicConfig(level=logging.INFO)

router = Router()

# Клавиатура выбора модели
model_keyboard = ReplyKeyboardMarkup(
    keyboard=[[KeyboardButton(text='GigaChat'), KeyboardButton(text='Yandex Cloud')]],
    resize_keyboard=True,
    one_time_keyboard=True
)

# В памяти будем хранить выбранную модель для каждого пользователя
user_models = {}

# --- Инициализация базы данных ---
def init_db():
    conn = sqlite3.connect('bot_database.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

# --- Инициализация Excel-файла ---
def init_excel():
    try:
        wb = openpyxl.load_workbook('users.xlsx')
        if 'Users' in wb.sheetnames:
            ws = wb['Users']
        else:
            ws = wb.create_sheet('Users')
            ws.append(['id', 'username', 'created_at'])
            wb.save('users.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('Users')
        ws.append(['id', 'username', 'created_at'])
        # Удаляем стандартный лист, если он есть
        if 'Sheet' in wb.sheetnames:
            std = wb['Sheet']
            wb.remove(std)
        wb.save('users.xlsx')

# --- Добавление пользователя в Excel ---
def add_user_excel(username):
    import os
    if not os.path.exists('users.xlsx'):
        init_excel()
    wb = openpyxl.load_workbook('users.xlsx')
    if 'Users' in wb.sheetnames:
        ws = wb['Users']
    else:
        ws = wb.create_sheet('Users')
        ws.append(['id', 'username', 'created_at'])
    # Проверяем, есть ли уже такой пользователь
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == username:
            return  # Уже есть
    new_id = ws.max_row  # id = номер строки (начиная с 1, но первая строка — заголовки)
    ws.append([new_id, username, datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    wb.save('users.xlsx')

# --- Модификация add_user для добавления в Excel ---
def add_user(username):
    conn = sqlite3.connect('bot_database.db')
    cursor = conn.cursor()
    try:
        cursor.execute('INSERT INTO users (username) VALUES (?)', (username,))
        conn.commit()
    except sqlite3.IntegrityError:
        pass  # Пользователь уже существует
    finally:
        conn.close()
    add_user_excel(username)

def get_user(username):
    conn = sqlite3.connect('bot_database.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM users WHERE username = ?', (username,))
    user = cursor.fetchone()
    conn.close()
    return user

# --- Вызов инициализации базы данных и Excel при запуске ---
init_db()
init_excel()

@router.message(F.text == "/start")
async def cmd_start(message: types.Message):
    user_id = message.from_user.id if message.from_user and message.from_user.id else 0
    user_models[user_id] = 'GigaChat'  # По умолчанию
    username = message.from_user.username if message.from_user and message.from_user.username else str(user_id)
    add_user(username)
    await message.answer(
        "Привет! Я бот с поддержкой GigaChat и Yandex Cloud.\n"
        "Выберите модель для общения:",
        reply_markup=model_keyboard
    )

@router.message(F.text == "/help")
async def cmd_help(message: types.Message):
    await message.answer(
        "/start — начать работу\n"
        "/help — показать все команды\n"
        "/model — выбрать модель (GigaChat или Yandex Cloud)\n"
        "Просто напишите сообщение, и я отвечу!"
    )

@router.message(F.text == "/model")
async def cmd_model(message: types.Message):
    await message.answer(
        "Выберите модель для общения:",
        reply_markup=model_keyboard
    )

@router.message(F.text.in_(["GigaChat", "Yandex Cloud"]))
async def choose_model(message: types.Message):
    user_id = message.from_user.id if message.from_user and message.from_user.id else 0
    user_models[user_id] = message.text
    await message.answer(f"Модель переключена на: {message.text}", reply_markup=ReplyKeyboardRemove())

@router.message()
async def handle_message(message: types.Message):
    user_id = message.from_user.id if message.from_user and message.from_user.id else 0
    model = user_models.get(user_id, 'GigaChat')
    user_text = message.text
    if model == 'GigaChat':
        # Здесь будет запрос к GigaChat API
        response = f"[GigaChat] Ответ на: {user_text}"
    else:
        # Здесь будет запрос к Yandex Cloud API
        response = f"[Yandex Cloud] Ответ на: {user_text}"
    await message.answer(response)

async def main():
    if not TELEGRAM_BOT_TOKEN:
        print('Ошибка: TELEGRAM_BOT_TOKEN не найден. Проверьте .env файл.')
        return
    bot = Bot(token=TELEGRAM_BOT_TOKEN)
    dp = Dispatcher()
    dp.include_router(router)
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main()) 